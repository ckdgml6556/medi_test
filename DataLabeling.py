import os
import re
import csv
import openpyxl
import pandas as pd
import Const

XML_READ_FILE_PATH = "data\ori_data.xlsx"
CSV_ALL_DATA = "data\stage_2_train.csv"
CSV_WRITE_FILE_PATH = "data\labeling_data.csv"
DATA_FILE_PATH = "data\\train\\"
#
# XML_READ_FILE_PATH = "data\stage_2_train.csv"
# XML_WRITE_FILE_PATH = "data\labeling_data_all.xlsx"

id_list = []
label_list = []


# def getClassificationLabel():
#     # data_only=True로 해야 수식이 아닌 값으로 불러옴
#     load_wb = openpyxl.load_workbook(XML_READ_FILE_PATH, data_only=True)
#     load_ws = load_wb['Sheet1']
#
#     index = 0
#     while True:
#         # 새로운 환자 데이터 id의 시작 index
#         item_index = 2 + (6 * index)
#         id = load_ws.cell(item_index, 1).value
#
#         if not id:
#             break
#         else:
#             id = id.split("_")[1]
#         label = ""
#         any = load_ws.cell(item_index, 2).value
#         print(f"save id = {id} any = {any}")
#
#         # 출혈이 없는 경우
#         if any == 0:
#             label = "nomal"
#             print(f"id = {id} label = {label}")
#         else:
#             # 어디선가 출혈 부위가 있는 경우 찾기
#             for i in range(1, 6):
#                 bleeder_index = 2 + (6 * index) + i
#                 part = load_ws.cell(bleeder_index, 1).value.split("_")[2]
#                 value = load_ws.cell(bleeder_index, 2).value
#                 if value == 1:
#                     label = part
#                     print(f"save id = {id} label = {label}")
#                     break
#         # 만약 전부 찾았는데 출혈 부위를 못찾은 경우
#         if not label:
#             print("not find label")
#         else:
#             id_list.append(id)
#             label_list.append(label)
#         index += 1


# Label CSV파일 기준으로 있는 데이터만 새로저장
def restoreData():
    csv_data = pd.read_csv(CSV_ALL_DATA, header=None)
    last_data_id = re.compile("_\w+.").search(os.listdir(Const.DATA_ALL_PATH).pop()).group()[1:11]

    label = ""
    index = 0
    try:
        while True:
            item_index = 1 + (6 * index)
            id = csv_data[0][item_index].split("_")[1]
            if os.path.isfile(f"{Const.DATA_ALL_PATH}ID_{id}.dcm"):
                print(f"Find File {id}")
                any = csv_data[1][item_index]
                print(any)
                if any == "0":
                    label = "nomal"
                    print(f"id = {id} label = {label}")
                else:
                    # 어디선가 출혈 부위가 있는 경우 찾기
                    for i in range(1, 6):
                        bleeder_index = 1 + (6 * index) + i
                        part = csv_data[0][bleeder_index].split("_")[2]
                        value = csv_data[1][bleeder_index]
                        if value == "1":
                            label = part
                            print(f"Bleeder id = {id} label = {label}")
                            break
                # 만약 전부 찾았는데 출혈 부위를 못찾은 경우
                if not label:
                    print("not find label")
                else:
                    id_list.append(id)
                    label_list.append(label)

            if last_data_id == id:
                print("FIND LAST ID")
                break
            index += 1
    except Exception as e:
        print(e)

def saveXlsx():
    if os.path.isfile(CSV_WRITE_FILE_PATH):
        os.remove(CSV_WRITE_FILE_PATH)
    label_df = pd.DataFrame({'id': id_list, 'label': label_list})
    print(label_df)
    label_df.to_csv(
        CSV_WRITE_FILE_PATH,
        index=False
    )

#전체데이터 sorting
def allDataCSVSorting():
    all_data = pd.read_csv(CSV_ALL_DATA)
    all_data = all_data.sort_values("ID", ascending=True)
    # all_data = pd.read_csv(CSV_ALL_DATA, header=None)
    print(all_data)
    all_data.to_csv(CSV_ALL_DATA,index=False)

#allDataCSVSorting()
restoreData()
saveXlsx()
# getClassificationLabel()
# saveXlsx()
